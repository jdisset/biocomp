### {{{                          --     import     --
from biocomp import utils as ut
import argparse
import json
import sys
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
##────────────────────────────────────────────────────────────────────────────}}}
### {{{                        --     CLIProgram     --

class CLIProgram:
    def __init__(self):
        self.is_notebook = 'ipykernel' in sys.modules
        self.parser = argparse.ArgumentParser()

    def add_argument(self, *args, **kwargs):
        self.parser.add_argument(*args, **kwargs)

    def parse_args(self, default_args=None):
        extra_args = default_args if default_args is not None else []

        # combine parsed args and extra_args. parsed args have priority over extra_args.
        # if we're in a notebook, only use extra_args. Otherwise we can combine them.
        if self.is_notebook:
            self.args = self.parser.parse_args(extra_args)
        else:
            self.args = self.parser.parse_args(extra_args + sys.argv[1:])
            ut.logger.info(f'args: {self.args}')

        self._postprocess_args()

    def _postprocess_args(self):
        pass

    def __getattr__(self, attr):
        if attr in self.__dict__:
            return self.__dict__[attr]
        elif hasattr(self, 'args') and hasattr(self.args, attr):
            return getattr(self.args, attr)
        else:
            raise AttributeError(f"{self.__class__.__name__} object has no attribute '{attr}'")

##────────────────────────────────────────────────────────────────────────────}}}
### {{{                         --     df tools     --
def merge_update(left_df, right_df, key_column, priority, use_left=None, use_right=None, how='outer'):
    """
    Merge two pandas dataframes with priority-based column selection.

    Parameters:
    priority (str): 'left' or 'right' to set priority dataframe for overlapping columns.
    use_left (list[str], optional): Columns to forcibly use from left_df.
    use_right (list[str], optional): Columns to forcibly use from right_df.

    Returns:
    pd.DataFrame: Merged dataframe based on the specified rules.
    """

    use_right = use_right or []
    use_left = use_left or []

    if not set(use_left).isdisjoint(use_right):
        raise ValueError("Columns in use_left and use_right must be disjoint")

    common_columns = set(left_df.columns).intersection(set(right_df.columns)).difference([key_column])

    # Rename common columns in right_df to avoid suffixes in the merged dataframe
    rename_columns = {col: col + '_right' for col in common_columns if col not in use_left}
    right_df_renamed = right_df.rename(columns=rename_columns)

    merged_df = pd.merge(left_df, right_df_renamed, on=key_column, how=how)

    # Apply use_left, use_right, and priority rules
    for col in common_columns:
        if col in use_right or (col + '_right' in merged_df and priority == 'right'):
            merged_df[col] = merged_df[col + '_right']

        merged_df.drop(columns=[col + '_right'], inplace=True, errors='ignore')

    return merged_df



def reorder_columns_front(df, columns):
    """
    Puts the specified columns in front of the dataframe.
    """
    columns = list(columns)
    return df[columns + [col for col in df.columns if col not in columns]]


def reorder_columns_back(df, columns):
    """
    Puts the specified columns in back of the dataframe.
    """
    columns = list(columns)
    return df[[col for col in df.columns if col not in columns] + columns]


##────────────────────────────────────────────────────────────────────────────}}}
### {{{                 --     xls manipulation     --

# def create_database_file(database_path, sheet_names):
    # print(f'creating database file {database_path}')
    # # create path if it doesn't exist
    # database_path = Path(database_path)
    # database_path.parent.mkdir(parents=True, exist_ok=True)
    # # create the database file
    # import pandas as pd
    # writer = pd.ExcelWriter(database_path, engine='openpyxl')
    # for sheet_name in sheet_names:
        # pd.DataFrame().to_excel(writer, sheet_name=sheet_name)
    # writer.close()

# def get_writer(database_path, create_if_not_exists=True):
    # database_path = Path(database_path)
    # if not database_path.exists():
        # if create_if_not_exists:
            # create_database_file(database_path, [])
        # else:
            # raise ValueError(f'database file {database_path} does not exist')
    # if database_path.suffix != '.xlsx':
        # raise ValueError(f'database file {database_path} must be an excel file')
    # return pd.ExcelWriter(database_path, engine='openpyxl')

# from openpyxl import load_workbook

# def create_sheet_if_not_exists(database_path, sheet_name):
    # database_path = Path(database_path)
    # if not database_path.exists():
        # raise ValueError(f'Database file {database_path} does not exist')
    # # Load the existing workbook

    # with open(database_path, 'rb') as f:
        # book = load_workbook(f)
        # # Check if the sheet exists
        # if sheet_name in book.sheetnames:
            # return
        # # Add new sheet
        # book.create_sheet(sheet_name)
        # # Save the workbook
        # book.save(database_path)


# def load_database_table(database_path, sheet_name, create_if_not_exists=False):
    # get_writer(database_path)
    # if create_if_not_exists:
        # create_sheet_if_not_exists(database_path, sheet_name)
    # return pd.read_excel(database_path, sheet_name=sheet_name, engine='openpyxl')

# def save_database_table(df, database_path, sheet_name):
    # # save the dataframe to the database at the specified sheet
    # # DO NOT overwrite the entire database file, just the specified sheet
    # writer = get_writer(database_path)
    # df.to_excel(writer, sheet_name=sheet_name, index=False)
    # writer.close()


import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def style_header_row(sheet, fg_color, bg_color, min_width=5, max_width=100):
    # Load the workbook and select the sheet

    # Apply styles to the header row
    for column in sheet.iter_cols(min_row=1, max_row=1):
        for cell in column:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.font = Font(color=fg_color)

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max(min_width, min(max_length + 2, max_width))
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


def wrap_text_all_cells(sheet):
    # Apply text wrap to all cells in the worksheet
    for row in sheet:
        for cell in row:
            cell.alignment = Alignment(wrapText=True)



from openpyxl import Workbook, load_workbook

def create_database_file(database_path, sheet_names):
    print(f'Creating database file {database_path}')
    database_path = Path(database_path)
    database_path.parent.mkdir(parents=True, exist_ok=True)
    
    workbook = Workbook()
    for sheet_name in sheet_names:
        workbook.create_sheet(title=sheet_name)
    # Remove default sheet
    del workbook['Sheet']
    workbook.save(filename=database_path)

def create_sheet_if_not_exists(database_path, sheet_name):
    database_path = Path(database_path)
    if not database_path.exists():
        raise ValueError(f'Database file {database_path} does not exist')

    workbook = load_workbook(filename=database_path)
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(title=sheet_name)
        workbook.save(filename=database_path)

def load_database_table(database_path, sheet_name, create_if_not_exists=False):
    database_path = Path(database_path)
    if not database_path.exists() or sheet_name not in load_workbook(filename=database_path).sheetnames:
        if create_if_not_exists:
            create_sheet_if_not_exists(database_path, sheet_name)
        else:
            raise ValueError(f'Sheet {sheet_name} does not exist in {database_path}')
    return pd.read_excel(database_path, sheet_name=sheet_name, engine='openpyxl')

def save_database_table(df, database_path, sheet_name):
    database_path = Path(database_path)
    if not database_path.exists():
        raise ValueError(f'Database file {database_path} does not exist')

    book = load_workbook(database_path)
    writer = pd.ExcelWriter(database_path, engine='openpyxl')
    writer.book = book
    writer.sheets.update({ws.title: ws for ws in book.worksheets})

    df.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.close()


##────────────────────────────────────────────────────────────────────────────}}}

